import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import datetime
import io
import re
import html
import joblib
import numpy as np
import nltk
from pathlib import Path
from difflib import SequenceMatcher
from collections import defaultdict

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dossier Intelligence · Procesador",
    page_icon="📰",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght=300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #F7F9F8; }
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 1100px; }

    .app-header {
        background: linear-gradient(135deg, #0D4F3C 0%, #1A7A5E 60%, #22A37A 100%);
        border-radius: 12px; padding: 1.1rem 1.8rem; margin-bottom: 1.2rem;
        position: relative; overflow: hidden; display: flex; align-items: center; gap: 1.2rem;
    }
    .app-header::before {
        content: ''; position: absolute; top: -30px; right: -30px;
        width: 120px; height: 120px; background: rgba(255,255,255,0.06); border-radius: 50%;
    }
    .app-header p { color: rgba(255,255,255,0.9); font-size: 0.92rem; font-weight: 500; margin: 0; }
    .app-header .badge {
        display: inline-block; background: rgba(255,255,255,0.15); color: #fff;
        font-size: 0.7rem; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase;
        padding: 0.2rem 0.6rem; border-radius: 20px; border: 1px solid rgba(255,255,255,0.25);
        white-space: nowrap; flex-shrink: 0;
    }

    .card {
        background: #FFFFFF; border: 1px solid #E8EFEC; border-radius: 12px;
        padding: 1.6rem 1.8rem; margin-bottom: 1.2rem;
        box-shadow: 0 1px 4px rgba(13,79,60,0.06);
    }
    .card-title {
        font-size: 0.8rem; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase;
        color: #1A7A5E; margin-bottom: 1rem;
    }

    .metrics-row { display: flex; gap: 1rem; margin: 1.5rem 0; }
    .metric-card {
        flex: 1; background: #FFFFFF; border: 1px solid #E8EFEC; border-radius: 12px;
        padding: 1.4rem 1.6rem; text-align: center; box-shadow: 0 1px 4px rgba(13,79,60,0.06);
    }
    .metric-card .metric-value {
        font-size: 2.2rem; font-weight: 700; color: #0D4F3C; line-height: 1;
        margin-bottom: 0.4rem; font-family: 'DM Mono', monospace;
    }
    .metric-card .metric-label { font-size: 0.78rem; color: #6B8F82; font-weight: 500; letter-spacing: 0.03em; }
    .metric-card.accent .metric-value { color: #1A7A5E; }
    .metric-card.muted  .metric-value { color: #9BB5AC; }

    .success-banner {
        background: linear-gradient(135deg, #0D4F3C, #1A7A5E); border-radius: 12px;
        padding: 1.2rem 1.8rem; display: flex; align-items: center; gap: 1rem;
        margin: 1.5rem 0; box-shadow: 0 2px 12px rgba(13,79,60,0.2);
    }
    .success-banner .icon { font-size: 1.6rem; line-height: 1; }
    .success-banner .text strong { display: block; color: #FFFFFF; font-size: 1rem; font-weight: 700; margin-bottom: 0.15rem; }
    .success-banner .text span { color: rgba(255,255,255,0.72); font-size: 0.85rem; }

    .step { display: flex; align-items: flex-start; gap: 1rem; margin-bottom: 0.8rem; }
    .step-num {
        min-width: 28px; height: 28px; background: #1A7A5E; color: white; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-size: 0.78rem; font-weight: 700; margin-top: 1px;
    }
    .step-text { color: #2D5A4A; font-size: 0.9rem; line-height: 1.6; }

    [data-testid="stFileUploader"] {
        background: #FFFFFF; border: 2px dashed #B2D4C8; border-radius: 12px;
        padding: 0.5rem; transition: border-color 0.2s;
    }
    [data-testid="stFileUploader"]:hover { border-color: #1A7A5E; }

    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0D4F3C, #1A7A5E); color: white;
        border: none; border-radius: 10px; padding: 0.7rem 2rem;
        font-family: 'DM Sans', sans-serif; font-size: 0.95rem; font-weight: 600;
        letter-spacing: 0.01em; transition: all 0.2s;
        box-shadow: 0 2px 8px rgba(13,79,60,0.25); width: 100%; height: 48px;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #0A3D2E, #156A50);
        box-shadow: 0 4px 14px rgba(13,79,60,0.35); transform: translateY(-1px);
    }
    .stButton > button[kind="primary"]:disabled { background: #C5D9D4; box-shadow: none; transform: none; }

    .stDownloadButton > button {
        background: #FFFFFF; color: #0D4F3C; border: 2px solid #1A7A5E;
        border-radius: 10px; padding: 0.65rem 1.8rem;
        font-family: 'DM Sans', sans-serif; font-size: 0.9rem; font-weight: 600;
        transition: all 0.2s; width: 100%; height: 48px;
    }
    .stDownloadButton > button:hover {
        background: #F0F7F4; box-shadow: 0 2px 10px rgba(13,79,60,0.15); transform: translateY(-1px);
    }

    .stWarning { background: #FFF8ED; border: 1px solid #F5C97A; border-radius: 10px; }
    .stError   { background: #FEF2F2; border: 1px solid #FCA5A5; border-radius: 10px; }
    .stProgress > div > div > div { background: linear-gradient(90deg, #1A7A5E, #22A37A); border-radius: 4px; }
    .streamlit-expanderHeader { background: #F7F9F8; border-radius: 8px; font-weight: 500; color: #0D4F3C; }
    [data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; border: 1px solid #E8EFEC; }
    hr { border: none; border-top: 1px solid #E8EFEC; margin: 1.5rem 0; }
    .stSpinner > div { border-top-color: #1A7A5E !important; }

    .file-status {
        display: flex; align-items: center; gap: 0.6rem; padding: 0.7rem 1rem;
        border-radius: 8px; font-size: 0.87rem; font-weight: 500; margin-top: 0.5rem;
    }
    .file-status.ok      { background: #F0F7F4; color: #0D4F3C; border: 1px solid #B2D4C8; }
    .file-status.missing { background: #FFF8ED; color: #92400E; border: 1px solid #F5C97A; }

    .results-header {
        font-size: 1.1rem; font-weight: 700; color: #0D4F3C;
        margin: 1.8rem 0 1rem 0; padding-bottom: 0.5rem; border-bottom: 2px solid #E8EFEC;
    }
</style>
""", unsafe_allow_html=True)

# NLTK stopwords
try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    with st.spinner("Descargando recursos de lenguaje por primera vez..."):
        nltk.download('stopwords')

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────
SIMILARITY_THRESHOLD_TITULOS = 0.93

FINAL_ORDER = [
    "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa",
    "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión",
    "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", "Tono", "Tema",
    "Temas Generales - Tema", "Resumen - Aclaracion", "Link Nota",
    "Link (Streaming - Imagen)", "Menciones - Empresa", "ID Original Retenido"
]

TIPO_MEDIO_MAP = {
    'online': 'Internet', 'internet': 'Internet',
    'diario': 'Prensa',
    'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
    'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
    'television': 'Televisión', 'televisión': 'Televisión',
    'revista': 'Revistas', 'revistas': 'Revistas',
}

# ──────────────────────────────────────────────────────────────────────────────
# CARGA DE MODELOS PKL
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def load_ml_models():
    try:
        sentiment_pipeline = joblib.load('pipeline_sentimiento.pkl')
        topic_pipeline     = joblib.load('pipeline_tema.pkl')
        return sentiment_pipeline, topic_pipeline
    except FileNotFoundError as e:
        st.error(
            f"**Error Crítico:** No se encontró `{e.filename}`. "
            "Asegúrate de que `pipeline_sentimiento.pkl` y `pipeline_tema.pkl` "
            "estén en la misma carpeta que esta app."
        )
        st.stop()

# ──────────────────────────────────────────────────────────────────────────────
# CARGA AUTOMÁTICA DESDE GOOGLE SHEETS
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_data
def load_config_from_sheets():
    """
    Obtiene las URLs publicadas en formato CSV desde st.secrets de Streamlit
    y construye los diccionarios de mapeo equivalentes a Configuracion.xlsx.
    """
    try:
        regiones_url = st.secrets["REGIONES_CSV_URL"]
        internet_url = st.secrets["INTERNET_CSV_URL"]
        menciones_url = st.secrets["MENCIONES_CSV_URL"]
        mapa_temas_url = st.secrets["MAPA_TEMAS_CSV_URL"]
    except KeyError as e:
        st.error(f"Falta configurar el secreto `{e.args[0]}` en los Secrets de Streamlit.")
        st.stop()

    # Lectura de los CSV remotos
    df_regiones = pd.read_csv(regiones_url)
    df_internet = pd.read_csv(internet_url)
    df_menciones = pd.read_csv(menciones_url)
    df_mapa_temas = pd.read_csv(mapa_temas_url)

    # Construcción de mapas
    region_map = pd.Series(
        df_regiones.iloc[:, 1].values,
        index=df_regiones.iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()

    internet_map = pd.Series(
        df_internet.iloc[:, 1].values,
        index=df_internet.iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()

    mention_map = {}
    for _, row in df_menciones.iterrows():
        key = row.iloc[0]
        val = row.iloc[1]
        if pd.notna(key) and pd.notna(val):
            mention_map[normalize_key(str(key))] = str(val).strip()

    final_topic_map = pd.Series(
        df_mapa_temas.iloc[:, 1].values,
        index=df_mapa_temas.iloc[:, 0].astype(str).str.strip()
    ).to_dict()

    return region_map, internet_map, mention_map, final_topic_map

# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES DE TEXTO Y NÚMEROS
# ──────────────────────────────────────────────────────────────────────────────
def normalize_key(s):
    if not isinstance(s, str):
        return ""
    s = html.unescape(s)
    s = re.sub(r'[\u2010-\u2015\u2212]', '-', s)
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()


def get_col(df, *names, default=''):
    normalized_cols = {normalize_key(c): c for c in df.columns}
    for name in names:
        if name in df.columns:
            return df[name]
        key = normalize_key(name)
        if key in normalized_cols:
            return df[normalized_cols[key]]
    return pd.Series([default] * len(df), index=df.index)


def convert_html_entities(text):
    if not isinstance(text, str):
        return text
    text = html.unescape(text)
    for bad, good in {
        '\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
        'Â': '', 'â': '', '€': '', '™': '', '\x9d': '', '\xa0': ' ',
    }.items():
        text = text.replace(bad, good)
    return text

def clean_text(text):
    if not isinstance(text, str):
        return text
    return convert_html_entities(text).strip()

def clean_title(title):
    if not isinstance(title, str):
        return ""
    return convert_html_entities(title)

def clean_title_for_output(title):
    if not isinstance(title, str):
        return ""
    title = convert_html_entities(title)
    title = title.replace('\n', ' ').replace('\r', ' ')
    title = re.sub(r'\s+\|.*$', '', title)
    title = re.sub(r'\|\s+.*$', '', title)
    title = re.sub(r'\s+-\s+.*$', '', title)
    return title.strip()

def normalize_title_for_comparison(title):
    if not isinstance(title, str):
        return ""
    cleaned_title = clean_title_for_output(title)
    abbreviations = {'tm': 'transporte masivo'}
    for abbr, full_text in abbreviations.items():
        cleaned_title = re.sub(fr'\b{abbr}\b', full_text, cleaned_title, flags=re.IGNORECASE)
    normalized_title = re.sub(r'\W+', ' ', cleaned_title).lower().strip()
    return normalized_title

def corregir_resumen(text):
    if not isinstance(text, str):
        return text
    text = convert_html_entities(text)
    text = re.sub(r'(<br\s*/?>|\[\.\.\.\]|\s+)', ' ', text).strip()
    m = re.search(r'[A-Z]', text)
    if m:
        text = text[m.start():]
    if text and not text.endswith('...'):
        text = text.rstrip('.') + '...'
    return text

def clean_cuerpo(text):
    if not isinstance(text, str) or not text.strip():
        return text
    text = convert_html_entities(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()

def preprocess_text_for_topic(text):
    if not isinstance(text, str):
        return ""
    try:
        from nltk.corpus import stopwords
        stop_words = set(stopwords.words('spanish'))
    except Exception:
        stop_words = set([
            'de','la','que','el','en','y','a','los','del','se','las','por',
            'un','para','con','no','una','su','al','lo'
        ])
    tokens = re.findall(r'\b\w+\b', text.lower())
    return " ".join(t for t in tokens if t not in stop_words)

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s:
        return None

    s = s.replace('$', '').replace(' ', '')

    if '.' not in s and ',' not in s:
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except ValueError:
            return None

    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        if s.count(',') > 1:
            s = s.replace(',', '')
        else:
            parts = s.split(',')
            if len(parts[-1]) == 3: 
                s = s.replace(',', '')
            else: 
                s = s.replace(',', '.')
    elif '.' in s:
        if s.count('.') > 1:
            s = s.replace('.', '')
        else:
            parts = s.split('.')
            if len(parts[-1]) == 3: 
                s = s.replace('.', '')
            else:
                pass

    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None

def normalize_url(url) -> str:
    if not isinstance(url, str):
        return ""
    url = url.strip().lower().rstrip('/')
    if not url.startswith('http'):
        return ""
    url = re.sub(r'^(https?://)www\.', r'\1', url)
    url = re.split(r'[?#]', url)[0].rstrip('/')
    return url

def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    return cell.value

# ──────────────────────────────────────────────────────────────────────────────
# LECTURA Y NORMALIZACIÓN DEL DOSSIER
# ──────────────────────────────────────────────────────────────────────────────
def read_and_normalize_dossier(wb_sheet, region_map, internet_map):
    headers = [cell.value for cell in wb_sheet[1] if cell.value is not None]
    raw_rows = []

    for row in wb_sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i >= len(row):
                row_data[h] = None
                continue
            cell = row[i]
            row_data[h] = extract_link_from_cell(cell)
        raw_rows.append(row_data)

    df = pd.DataFrame(raw_rows)

    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(TIPO_MEDIO_MAP)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av       = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica  = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'
    is_print    = df['Tipo de Medio'].isin(['Prensa', 'Revistas'])
    is_bcast    = df['Tipo de Medio'].isin(['Radio', 'Televisión'])

    if 'Medio' in df.columns:
        df['Región'] = (
            df['Medio'].astype(str).str.lower().str.strip()
            .map(region_map)
            .fillna('N/A')
        )
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia']         = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha']              = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora']               = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)

    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_title)

    df['Autor - Conductor']       = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina']             = df.get('Nro. Pagina', pd.Series(dtype=str))

    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión']               = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión']                 = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av      = get_col(df, 'CPE', default=np.nan)
    cpe_grafica = get_col(df, 'Valor de Nota', default=np.nan)
    df['CPE']   = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier']     = df.get('Tier',     pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono']     = ''
    df['Tema']     = ''
    df['Temas Generales - Tema'] = ''

    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)
    df['Resumen - Aclaracion'] = cuerpo_cleaned.apply(corregir_resumen)

    url_nota_av    = get_col(df, 'URL Nota AV', 'Link Nota AV', default='')
    url_streaming  = get_col(df, 'URL (Streaming - Imagen)', default='')
    url_nota_raw   = get_col(df, 'Link Nota', default='')

    link_nota_final    = []
    link_stream_final  = []

    for val_av, val_str, val_url_nota, av_row, int_row, print_row, bcast_row in zip(
        url_nota_av, url_streaming, url_nota_raw, is_av, is_internet, is_print, is_bcast
    ):
        raw_val_av = str(val_av).strip() if pd.notna(val_av) else ""
        raw_val_str = str(val_str).strip() if pd.notna(val_str) else ""
        raw_val_url_nota = str(val_url_nota).strip() if pd.notna(val_url_nota) else ""

        if av_row:
            if raw_val_av and "news.globalnews.com.ar" in raw_val_av:
                raw_val_av = raw_val_av.replace("news.globalnews.com.ar", "news.globalnews.com.co")
            link_nota_final.append(raw_val_av if raw_val_av else None)
        else:
            link_nota_final.append(raw_val_str if raw_val_str else None)

        if int_row:
            link_stream_final.append(raw_val_url_nota if raw_val_url_nota else None)
        else:
            link_stream_final.append(None)

    df['Link Nota']                = link_nota_final
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av      = get_col(df, 'Menciones - Empresa', default='').fillna('').astype(str).apply(clean_text)
    menciones_grafica = get_col(df, 'Empresa rel.', 'Empresa Rel.', 'Empresa Relacionada', default='').fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    return df

# ──────────────────────────────────────────────────────────────────────────────
# EXPANSIÓN POR ; EN MENCIONES
# ──────────────────────────────────────────────────────────────────────────────
def expand_menciones(df):
    rows_expanded = []
    for _, row in df.iterrows():
        menciones = [m.strip() for m in str(row['Menciones - Empresa']).split(';') if m.strip()]
        if not menciones:
            d = row.to_dict()
            d['Menciones - Empresa'] = ''
            rows_expanded.append(d)
        else:
            for m in menciones:
                d = row.to_dict()
                d['Menciones - Empresa'] = m
                rows_expanded.append(d)
    return pd.DataFrame(rows_expanded).reset_index(drop=True)

# ──────────────────────────────────────────────────────────────────────────────
# MAPEO DE MENCIONES
# ──────────────────────────────────────────────────────────────────────────────
def apply_mention_map(df, mention_map):
    if 'Menciones - Empresa' in df.columns:
        df['Menciones - Empresa'] = df['Menciones - Empresa'].apply(
            lambda x: mention_map.get(normalize_key(str(x)), str(x).strip()) if pd.notna(x) and str(x).strip() != '' else x
        )
    return df

# ──────────────────────────────────────────────────────────────────────────────
# DETECCIÓN DE DUPLICADOS (CON RASTREO DE ID ORIGINAL)
# ──────────────────────────────────────────────────────────────────────────────
def calculate_title_quality_score(title: str) -> int:
    if not isinstance(title, str):
        return -999
    score = 100
    score -= len(re.findall(r'&[#\w]+;', title)) * 10
    score -= title.count('??') * 5
    score -= title.count('\x9d') * 5
    if len(title) > 250: score -= 5
    if len(title) < 15:  score -= 20
    if '\n' in title:    score -= 15
    if '|' in title:     score -= 5
    return int(score)

def are_duplicates(row1: pd.Series, row2: pd.Series, title_similarity_threshold=0.93, date_proximity_days=1) -> bool:
    menc1 = str(row1.get('Menciones - Empresa', '')).strip().lower()
    menc2 = str(row2.get('Menciones - Empresa', '')).strip().lower()
    if menc1 != menc2:
        return False

    medio1 = str(row1.get('Medio', '')).strip().lower()
    medio2 = str(row2.get('Medio', '')).strip().lower()
    if medio1 != medio2:
        return False

    tipo1 = str(row1.get('Tipo de Medio', '')).strip()
    tipo2 = str(row2.get('Tipo de Medio', '')).strip()
    if tipo1 != tipo2:
        return False

    url_col = 'Link (Streaming - Imagen)'
    url1 = normalize_url(row1.get(url_col, ''))
    url2 = normalize_url(row2.get(url_col, ''))
    if url1 and url2 and url1 != url2:
        return False

    fecha1 = row1.get('Fecha')
    fecha2 = row2.get('Fecha')
    if pd.isna(fecha1) or pd.isna(fecha2):
        return False

    t1 = pd.to_datetime(fecha1)
    t2 = pd.to_datetime(fecha2)

    if tipo1 == 'Internet':
        if abs((t1 - t2).days) > date_proximity_days:
            return False
    else:
        if t1.date() != t2.date():
            return False

    if tipo1 in ['Radio', 'Televisión']:
        hora1 = str(row1.get('Hora', '')).strip()
        hora2 = str(row2.get('Hora', '')).strip()
        if hora1 and hora2 and hora1 != hora2:
            return False

    if tipo1 == 'Internet':
        link_nota_1 = normalize_url(row1.get('Link Nota', ''))
        link_nota_2 = normalize_url(row2.get('Link Nota', ''))
        if link_nota_1 and link_nota_2 and link_nota_1 == link_nota_2:
            return True

    titulo1 = normalize_title_for_comparison(row1.get('Título', ''))
    titulo2 = normalize_title_for_comparison(row2.get('Título', ''))

    if not titulo1 or not titulo2:
        return False

    if titulo1 == titulo2:
        return True

    if len(titulo1) > 15 and len(titulo2) > 15:
        if titulo1 in titulo2 or titulo2 in titulo1:
            return True

    similarity = SequenceMatcher(None, titulo1, titulo2).ratio()
    if similarity >= title_similarity_threshold:
        return True

    return False

def detect_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().reset_index(drop=True)
    df['original_index'] = df.index
    df['title_quality']  = df['Título'].apply(calculate_title_quality_score)

    df.sort_values(
        by=['title_quality', 'Fecha', 'original_index'],
        ascending=[False, True, True],
        inplace=True,
        na_position='last'
    )

    duplicate_indices = set()
    duplicate_to_original_id = {}
    grouping_keys     = ['Medio', 'Menciones - Empresa']

    for _, group in df.groupby(grouping_keys, dropna=False):
        if len(group) < 2:
            continue

        group_rows = group.to_dict('records')

        for i in range(len(group_rows)):
            current = group_rows[i]
            current_orig_idx = current['original_index']
            if current_orig_idx in duplicate_indices:
                continue

            for j in range(i + 1, len(group_rows)):
                compare = group_rows[j]
                compare_orig_idx = compare['original_index']
                if compare_orig_idx in duplicate_indices:
                    continue

                if are_duplicates(pd.Series(current), pd.Series(compare), title_similarity_threshold=SIMILARITY_THRESHOLD_TITULOS):
                    duplicate_indices.add(compare_orig_idx)
                    duplicate_to_original_id[compare_orig_idx] = current.get('ID Noticia')

    df['is_duplicate'] = df['original_index'].isin(duplicate_indices)
    df['ID Original Retenido'] = df['original_index'].map(duplicate_to_original_id)

    df.sort_values('original_index', inplace=True)
    df.drop(columns=['original_index', 'title_quality'], inplace=True)
    return df.reset_index(drop=True)

# ──────────────────────────────────────────────────────────────────────────────
# CLASIFICACIÓN PKL (Tono + Tema)
# ──────────────────────────────────────────────────────────────────────────────
LABEL_MAP_TONO = {1: 'Positivo', 0: 'Neutro', -1: 'Negativo',
                  '1': 'Positivo', '0': 'Neutro', '-1': 'Negativo'}

def classify_with_pkl(df, sentiment_pipeline, topic_pipeline, final_topic_map, progress_bar):
    df = df.copy()
    mask_unique = ~df['is_duplicate']
    df_valid = df[mask_unique].copy()

    if df_valid.empty:
        df['Tono'] = ''
        df['Temas Generales - Tema'] = ''
        df['Tema'] = ''
        return df

    df_valid['_texto_ia'] = (
        df_valid['Título'].fillna('') + ' ' + df_valid['Resumen - Aclaracion'].fillna('')
    )

    progress_bar.progress(60, text="Paso 5 / 7 — Clasificando tono con modelo PKL...")
    preds_sent = sentiment_pipeline.predict(df_valid['_texto_ia'])
    df_valid['Tono'] = [LABEL_MAP_TONO.get(p, str(p).title()) for p in preds_sent]

    progress_bar.progress(72, text="Paso 6 / 7 — Clasificando temas con modelo PKL...")
    df_valid['_resumen_procesado'] = df_valid['_texto_ia'].apply(preprocess_text_for_topic)
    df_valid['Temas Generales - Tema'] = topic_pipeline.predict(df_valid['_resumen_procesado'])

    df_valid['_titulo_norm'] = df_valid['Título'].apply(normalize_title_for_comparison)
    homo = df_valid.groupby('_titulo_norm')['Temas Generales - Tema'].transform(
        lambda x: x.mode()[0] if not x.mode().empty else x
    )
    df_valid['Temas Generales - Tema'] = homo

    df_valid['Tema'] = (
        df_valid['Temas Generales - Tema'].astype(str).str.strip()
        .map(final_topic_map)
        .fillna('Indefinido')
    )

    df_valid.drop(columns=['_texto_ia', '_resumen_procesado', '_titulo_norm'], inplace=True)

    df.update(df_valid[['Tono', 'Temas Generales - Tema', 'Tema']])

    mask_dup = df['is_duplicate']
    if mask_dup.any():
        df.loc[mask_dup, 'Tono']                    = 'Duplicada'
        df.loc[mask_dup, 'Temas Generales - Tema']  = '-'
        df.loc[mask_dup, 'Tema']                    = '-'

    return df

# ──────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL EXCEL DE SALIDA
# ──────────────────────────────────────────────────────────────────────────────
def to_excel(df):
    cols   = [c for c in FINAL_ORDER if c in df.columns]
    df_out = df[cols].copy()

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Resultado'

    font_link     = Font(color='000000', underline=None)
    font_header   = Font(bold=True)
    align_left    = Alignment(horizontal='left')
    date_fmt      = 'DD/MM/YYYY'
    fmt_thousands = '#,##0'
    fmt_currency  = '"$"#,##0'

    NUM_COLS = {
        'ID Noticia', 'Nro. Pagina', 'Dimensión', 'Duración - Nro. Caracteres',
        'CPE', 'Tier', 'Audiencia', 'ID Original Retenido'
    }

    ws.append(cols)
    for i, col_name in enumerate(cols, start=1):
        ws.cell(row=1, column=i).font = font_header

    for _, row in df_out.iterrows():
        out_row = []
        link_map = {}

        for ci, col_name in enumerate(cols, start=1):
            val = row[col_name]

            if col_name == 'Fecha':
                if pd.notna(val):
                    cv = val.to_pydatetime() if isinstance(val, pd.Timestamp) else val
                else:
                    cv = None

            elif col_name in NUM_COLS:
                cv = parse_numeric(val)

            elif col_name in ('Link Nota', 'Link (Streaming - Imagen)'):
                if pd.notna(val) and isinstance(val, str) and val.startswith('http'):
                    cv = 'Link'
                    link_map[ci] = val
                else:
                    cv = val if pd.notna(val) else None

            else:
                cv = None if (isinstance(val, float) and np.isnan(val)) else val
                if cv is not None and not isinstance(cv, (int, float, datetime.datetime, datetime.date)):
                    cv = str(cv)

            out_row.append(cv)

        ws.append(out_row)
        cur_row = ws.max_row

        for ci, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=cur_row, column=ci)

            if col_name == 'Fecha':
                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                    cell.number_format = date_fmt

            elif cell.value is not None and isinstance(cell.value, (int, float)):
                if col_name in ('Nro. Pagina', 'Dimensión', 'Duración - Nro. Caracteres', 'Tier', 'Audiencia'):
                    cell.number_format = fmt_thousands
                elif col_name == 'CPE':
                    cell.number_format = fmt_currency

        for ci, url in link_map.items():
            cell = ws.cell(row=cur_row, column=ci)
            cell.hyperlink = url
            cell.font      = font_link
            cell.alignment = align_left

    for i, col_name in enumerate(cols, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ('Título', 'Resumen - Aclaracion'):
            ws.column_dimensions[letter].width = 50
        elif 'Link' in col_name:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ──────────────────────────────────────────────────────────────────────────────

for _k in ("result_1", "result_2"):
    if _k not in st.session_state:
        st.session_state[_k] = None

st.markdown("""
<div class="app-header">
    <div class="badge">Media Intelligence · PKL</div>
    <p>Limpieza, normalización y clasificación automática de dossiers · v1.0</p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="card">
    <div class="card-title">Cómo usar esta herramienta</div>
    <div class="step">
        <div class="step-num">1</div>
        <div class="step-text">Asegúrate de configurar las URLs públicas CSV de tu Google Sheets en los <strong>Secrets de Streamlit</strong>,
        y que <strong>pipeline_sentimiento.pkl</strong> y <strong>pipeline_tema.pkl</strong> estén en la raíz.</div>
    </div>
    <div class="step">
        <div class="step-num">2</div>
        <div class="step-text">Sube uno o dos dossiers. Si subes dos, cada uno
        se procesa de forma independiente y obtienes un archivo descargable por separado.</div>
    </div>
    <div class="step">
        <div class="step-num">3</div>
        <div class="step-text">Haz clic en <strong>Iniciar proceso</strong>.
        Los botones de descarga aparecerán al finalizar.</div>
    </div>
</div>
""", unsafe_allow_html=True)

with st.expander("📋  Ver estructura de Secrets requerida (Google Sheets en CSV)"):
    st.markdown("""
    Debes agregar las siguientes claves en tu archivo local `.streamlit/secrets.toml` o en la consola de Streamlit Cloud:
    ```toml
    REGIONES_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-xxxx/pub?gid=0&single=true&output=csv"
    INTERNET_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-xxxx/pub?gid=123456&single=true&output=csv"
    MENCIONES_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-xxxx/pub?gid=789101&single=true&output=csv"
    MAPA_TEMAS_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-xxxx/pub?gid=112131&single=true&output=csv"
    ```
    
    *Cada hoja debe contar con 2 columnas (A y B) con el mismo orden del archivo original:*
    - **Regiones:** Medio | Región
    - **Internet:** Medio Original | Medio Mapeado
    - **Menciones:** Mención Original | Mención Mapeada
    - **Mapa_Temas:** Temas Generales - Tema | Tema
    """)

st.markdown("<br>", unsafe_allow_html=True)

st.markdown('<div class="card-title">Carga de Dossiers</div>', unsafe_allow_html=True)

col_up1, col_up2 = st.columns(2)

with col_up1:
    st.markdown("**Dossier 1 · Marca**")
    file_1 = st.file_uploader(
        "Dossier Marca",
        type=["xlsx"],
        accept_multiple_files=False,
        label_visibility="collapsed",
        key="uploader_1",
    )
    if file_1:
        st.markdown(
            f'<div class="file-status ok">✓ {file_1.name}</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="file-status missing">⚠ Sin archivo</div>',
            unsafe_allow_html=True,
        )

with col_up2:
    st.markdown("**Dossier 2 · Competencia** *(opcional)*")
    file_2 = st.file_uploader(
        "Dossier Competencia",
        type=["xlsx"],
        accept_multiple_files=False,
        label_visibility="collapsed",
        key="uploader_2",
    )
    if file_2:
        st.markdown(
            f'<div class="file-status ok">✓ {file_2.name}</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="file-status missing">⚠ Sin archivo</div>',
            unsafe_allow_html=True,
        )

st.markdown("<br>", unsafe_allow_html=True)

files_loaded = [(file_1, "Marca", "result_1"), (file_2, "Competencia", "result_2")]
files_loaded = [(f, lbl, key) for f, lbl, key in files_loaded if f is not None]
n_files = len(files_loaded)

start_clicked = st.button(
    f"▶  Procesar {n_files} dossier{'s' if n_files != 1 else ''}",
    disabled=n_files == 0,
    type="primary",
)

download_area = st.empty()

if start_clicked and files_loaded:
    st.session_state["result_1"] = None
    st.session_state["result_2"] = None

    for f, label, state_key in files_loaded:
        st.markdown(f"### {label} — `{f.name}`")
        st.markdown("<hr>", unsafe_allow_html=True)

        progress_bar = st.progress(0, text="Iniciando proceso...")

        # 1. Carga de recursos (Modelos y Configuración de Google Sheets)
        progress_bar.progress(5, text="Paso 1 / 7 — Cargando modelos PKL y configuración desde Google Sheets...")
        sentiment_pipeline, topic_pipeline = load_ml_models()

        try:
            region_map, internet_map, mention_map, final_topic_map = load_config_from_sheets()
        except Exception as e:
            st.error(f"**Error al cargar las configuraciones remotas de Google Sheets:** {e}")
            st.info("Revisa si tus URLs en secrets son válidas, públicas y de formato CSV.")
            continue

        # 2. Carga y normalización
        progress_bar.progress(15, text="Paso 2 / 7 — Leyendo Dossier y normalizando estructura...")
        wb = load_workbook(f, data_only=True)
        df = read_and_normalize_dossier(wb.active, region_map, internet_map)

        if df["Fecha"].isna().any():
            st.warning("⚠️ Algunas fechas no se pudieron convertir. Revisa el archivo original.")

        # 3. Menciones
        progress_bar.progress(28, text="Paso 3 / 7 — Expandiendo filas por menciones (;)...")
        df = expand_menciones(df)

        progress_bar.progress(35, text="Paso 4 / 7 — Aplicando mapeos de menciones...")
        df = apply_mention_map(df, mention_map)

        # 4. Duplicados
        progress_bar.progress(45, text="Paso 4b / 7 — Detectando duplicados de forma optimizada...")
        df = detect_duplicates(df)

        # 5. Inteligencia Artificial
        df = classify_with_pkl(df, sentiment_pipeline, topic_pipeline, final_topic_map, progress_bar)

        # 6. Salida
        progress_bar.progress(95, text="Paso 7 / 7 — Generando archivo de salida...")
        excel_data   = to_excel(df)
        filename     = f"Dossier_{label}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        total        = len(df)
        dups_count   = int(df["is_duplicate"].sum())
        unique_count = total - dups_count

        progress_bar.progress(100, text="✓ Proceso completado")

        st.session_state[state_key] = {
            "data":     excel_data,
            "filename": filename,
            "label":    label,
            "total":    total,
            "unique":   unique_count,
            "dups":     dups_count,
        }

        st.markdown(f"""
        <div class="success-banner">
            <div class="icon">⚡</div>
            <div class="text">
                <strong>{label} — Proceso finalizado correctamente</strong>
                <span>{total:,} filas · {unique_count:,} únicas · {dups_count:,} duplicadas</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="metrics-row">
            <div class="metric-card">
                <div class="metric-value">{total:,}</div>
                <div class="metric-label">Filas totales</div>
            </div>
            <div class="metric-card accent">
                <div class="metric-value">{unique_count:,}</div>
                <div class="metric-label">Noticias únicas</div>
            </div>
            <div class="metric-card muted">
                <div class="metric-value">{dups_count:,}</div>
                <div class="metric-label">Duplicadas</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

res1 = st.session_state.get("result_1")
res2 = st.session_state.get("result_2")

if res1 or res2:
    with download_area.container():
        st.markdown('<div class="card-title">Descargas</div>', unsafe_allow_html=True)
        dl_col1, dl_col2 = st.columns(2)

        with dl_col1:
            if res1:
                st.download_button(
                    label=f"⬇ Descargar {res1['label']}",
                    data=res1["data"],
                    file_name=res1["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_final_1",
                )

        with dl_col2:
            if res2:
                st.download_button(
                    label=f"⬇ Descargar {res2['label']}",
                    data=res2["data"],
                    file_name=res2["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_final_2",
                )
